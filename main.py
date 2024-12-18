import os
import re
import sys
from datetime import datetime

import fitz # pymupdf
from rich import print
from openpyxl import Workbook
from openpyxl.styles import Font
from  openpyxl import load_workbook

import web_scrape.scrape as scrape
import utils.ler_arquivos as leitura
from utils.eventos import registrar_evento
from version.metadata import print_project_information


def main():
    os.system('cls')
    print_project_information()
    app_dir = os.path.dirname(os.path.abspath(__file__))
    if not verificar_arquivos(app_dir):
        return False
    
    # Lista os números dos processos para leitura
    file_path = os.path.join(app_dir, 'arquivos', 'processos.txt')    
    processos = leitura.ler_txt(file_path, True)
    
    # Lista os números dos captchas
    captcha_path = os.path.join(app_dir, 'arquivos', 'lista_captcha.xlsx')
    captchas = leitura.ler_xlsx(captcha_path, 'lista', 'num')

    # Lista as entidades que das unidades que serão baixadas
    entidades_path = os.path.join(app_dir, 'arquivos', 'lista_entidades.xlsx')
    cod_entidades = leitura.ler_xlsx(entidades_path, 'dados', 'Código', 'Leitura', 'sim')

    print('\n[bold cyan] -> Baixando arquivos PDFs\n')
    # Realiza o download dos arquivos .PDFs
    scrape.main(app_dir, cod_entidades, captchas)
    
    print('[bold cyan] -> Descompactando arquivos\n')
    # Descompacta e lê cada arquivo .PDF
    descompacta_arquivos(app_dir)

    print('[bold cyan] -> Lendo arquivos baixados\n')
    ler_dados_pdf(app_dir, processos)

    print('[bold blue]***** FIM DO PROCESSO *****')


def descompacta_arquivos(app_dir:str)->list:
    # Pastas de origem e destino
    temp_path = os.path.join(app_dir, 'temp')
    downloads_path = os.path.expanduser("~/Downloads")
    os.makedirs(temp_path, exist_ok=True)
    today_date = datetime.now().date()

    # Verifica todos os arquivos na pasta de downloads
    for file_name in os.listdir(downloads_path):
        if file_name.lower().endswith(".zip"):  # Filtra apenas arquivos .zip
            file_path = os.path.join(downloads_path, file_name)
            modification_time = os.path.getmtime(file_path)  # Tempo de modificação
            modification_datetime = datetime.fromtimestamp(modification_time)  # Converte para datetime

            # Verifica se o arquivo foi modificado hoje
            if modification_datetime.date() == today_date:
                leitura.descompactar_zip(file_path, temp_path)


def ler_dados_pdf(app_dir: str, processos: list):
    files_dir = os.path.join(app_dir, 'temp')
    if not os.path.exists(files_dir) or not processos:
        registrar_evento()
        return False

    pdfs_files = [file for file in os.listdir(files_dir) if file.lower().endswith('.pdf')]
    
    for i, nome_pdf in enumerate(pdfs_files):
        pdf_path = os.path.join(files_dir, nome_pdf)
        com_devedora = None
        
        print(f'[bold green]  Lendo arquivo {i+1}/{len(pdfs_files)} -> {nome_pdf}')
        with fitz.open(pdf_path) as pdf:
            pagina = 0
            precatorio = dict()
            
            for index_page, page in enumerate(pdf):
                texto_pagina = page.get_text("text").split('\n')  # Extrai o texto da página divido em lista
                
                for linha_pagina in texto_pagina:
                    texto = linha_pagina.strip()
                    chave_processo = re.sub('\D+', '', texto)
                    if chave_processo in processos:
                        precatorio['num_autos'] = texto
                        pagina = index_page + 1
                    
                    if precatorio:
                        # Adiciona os dados ao arquivo excel
                        if len(precatorio) == 7:
                            precatorio['arquivo'] = nome_pdf
                            precatorio['pagina'] = pagina
                            escrever_dados(app_dir, precatorio)
                            precatorio.clear() # Limpa o dicionário após salvar em excel

                        # Processo
                        if not chave_processo in processos and re.match('\d{7}\-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}', texto):
                            precatorio['num_processo'] = texto.strip()
                        
                        # Natureza
                        if re.match('^(alimentar)|^(outras\s*esp[ée]cies)$', texto, re.I):
                            precatorio['natureza'] = texto

                        # Ordem orçamentária
                        if re.match('^(\d+\/\d{4})$', texto):
                            precatorio['ordem_orcamentaria'] = texto
                        
                        # Suspenso?                      
                        if re.match('^(suspenso)\W+\s([ns])$', texto, re.I):
                            suspenso = texto.split()
                            precatorio['suspenso'] = suspenso[1]
                        
                        # Data protocolo
                        if re.match('^(\d{2}\/\d{2}\/\d{4})', texto):
                            data_protocolo = re.search('(\d{2}\/\d{2}\/\d{4})', texto)
                            data_objeto = datetime.strptime(data_protocolo.group(), r'%d/%m/%Y')
                            precatorio['data_protocolo'] = data_objeto             
                        
                        # Devedora
                        if com_devedora:
                            precatorio['devedora'] = texto
                            com_devedora = None
                        com_devedora = re.match('^(devedora\:)$', texto, re.I)
    

def verificar_arquivos(app_dir:str):
    nomes_pastas = ['arquivos', 'logs', 'temp']
    [os.makedirs(os.path.join(app_dir, n), exist_ok=True) for n in nomes_pastas]

    nomes_arquivos = ['lista_captcha.xlsx', 'lista_entidades.xlsx', 'processos.txt']
    com_arquivos = True
    for n in nomes_arquivos:
        file_path = os.path.join(app_dir, 'arquivos', n)
        if not os.path.exists(file_path):
            registrar_evento(f'"{n}" não existe')
            com_arquivos = False

    return com_arquivos


def escrever_dados(app_dir:str, precatorio:dict):
    # criar o arquivo na primeira chamada
    file_path = os.path.join(app_dir, f'Relatório_{str(datetime.now().date())}.xlsx')
    if not os.path.exists(file_path):
        file_path = criar_arquivo_excel(app_dir)
    
    em_uso = True
    while em_uso:
        try:
            # Tentativa de abrir o arquivo em modo exclusivo
            with open(file_path, 'r+'):
                em_uso = False
        except IOError:
            print('Feche o Arquivo para eu poder salvar novos dados')
            sys.stdout.write("\033[F")  # Volta uma linha
            sys.stdout.write("\033[K")  # Limpa a linha atual

    # Carregar o arquivo existente
    workbook = load_workbook(file_path)
    # Verificar se a aba já existe
    if not 'Dados' in workbook.sheetnames:
        workbook.create_sheet(title='Dados')
    sheet = workbook['Dados']

    # Dados para adicionar
    data_formatada = precatorio.get('data_protocolo')
    new_data = [
        precatorio.get('num_autos'),
        precatorio.get('num_processo'),
        precatorio.get('natureza'),
        precatorio.get('ordem_orcamentaria'),
        precatorio.get('suspenso'),
        # precatorio.get('data_protocolo'),
        data_formatada.strftime("%d/%m/%Y") if data_formatada else '',
        precatorio.get('devedora'),
        precatorio.get('arquivo'),
        precatorio.get('pagina'),
    ]
    
    # Identificar a próxima linha disponível
    next_row = sheet.max_row + 1

    # Adicionar os dados na próxima linha
    for col_index, value in enumerate(new_data, start=1):
        sheet.cell(row=next_row, column=col_index, value=value)

    # Salvar as alterações
    workbook.save(file_path)


def criar_arquivo_excel(app_dir:str, num_file:str=''):
    file_path = os.path.join(app_dir, f'Relatório_{str(datetime.now().date())}{num_file}.xlsx')
    
    # Criar o workbook e a planilha
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Dados'  # Nome da planilha

    # Definir os rótulos das colunas
    headers = ['Autos', 'Processo', 'Natureza', 'Ordem orçamentária', 'Suspenso?', 'Data do Protocolo', 'Devedora', 'Arquivo', 'Página']

    # Adicionar os rótulos com formatação em negrito
    bold_font = Font(bold=True)
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = bold_font

    # Definir largura específica para as colunas
    column_widths = [24, 24, 17, 20, 9, 14, 60, 37]  # Largura das colunas (ajuste conforme necessário)
    for col_index, width in enumerate(column_widths, start=1):
        column_letter = sheet.cell(row=1, column=col_index).column_letter
        sheet.column_dimensions[column_letter].width = width

    # Aplicar auto filtro aos rótulos
    sheet.auto_filter.ref = f"A1:I1"  # Ajuste o intervalo conforme o número de colunas

    # Congelar a primeira linha
    sheet.freeze_panes = "A2"  # Congela a linha acima da célula especificada (neste caso, a linha 1)

    # Salvar o arquivo Excel
    workbook.save(file_path)
    return file_path


main()